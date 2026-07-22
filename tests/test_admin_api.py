"""api/routers/admin.py — HTTP-контракт веб-панели: доступ только админам,
формат списков (JSON) и Excel-экспорта (добавится в Task 7)."""
from __future__ import annotations

import dataclasses
import io
import time

import openpyxl
from fastapi.testclient import TestClient

from api.app import create_app
from conftest import _sign, _sqlite_lifecycle, _test_config, init_data_for


def _client(tg_id: int) -> tuple[TestClient, dict]:
    app = create_app(bot=object(), config=_test_config(), bot_lifecycle=_sqlite_lifecycle)
    return TestClient(app), {"X-Telegram-Init-Data": init_data_for(tg_id)}


def _admin_client(tg_id: int) -> tuple[TestClient, dict]:
    config = dataclasses.replace(_test_config(), owner_chat_id=tg_id)
    app = create_app(bot=object(), config=config, bot_lifecycle=_sqlite_lifecycle)
    return TestClient(app), {"X-Telegram-Init-Data": init_data_for(tg_id)}


def _admin_client_with_username(tg_id: int, username: str) -> tuple[TestClient, dict]:
    # init_data_for() only encodes {"id": tg_id} — no username. To exercise the
    # real `user.username if user else None` mapping in admin.py with a
    # non-None value, sign our own initData (same _sign() helper conftest
    # uses) that also carries a username, which current_client() then persists
    # onto the User row via touch_activity().
    config = dataclasses.replace(_test_config(), owner_chat_id=tg_id)
    app = create_app(bot=object(), config=config, bot_lifecycle=_sqlite_lifecycle)
    init_data = _sign({
        "auth_date": str(int(time.time())),
        "user": f'{{"id": {tg_id}, "username": "{username}"}}',
    })
    return TestClient(app), {"X-Telegram-Init-Data": init_data}


def _admin_client_with_first_name(tg_id: int, first_name: str) -> tuple[TestClient, dict]:
    # Same technique as _admin_client_with_username(), but for first_name.
    # Telegram usernames are restricted to [A-Za-z0-9_], but first_name has NO
    # such restriction -- a real Telegram user can set it to a formula-injection
    # payload (e.g. "=SUM(...)"). Picking a payload with no embedded quotes so
    # it survives this helper's naive (non-json.dumps) f-string interpolation
    # into the signed initData "user" field, same as the username helper does.
    config = dataclasses.replace(_test_config(), owner_chat_id=tg_id)
    app = create_app(bot=object(), config=config, bot_lifecycle=_sqlite_lifecycle)
    init_data = _sign({
        "auth_date": str(int(time.time())),
        "user": f'{{"id": {tg_id}, "first_name": "{first_name}"}}',
    })
    return TestClient(app), {"X-Telegram-Init-Data": init_data}


def test_leads_rejected_for_non_admin():
    client, headers = _client(801)
    with client:
        response = client.get("/api/admin/leads", headers=headers)
    assert response.status_code == 403


def test_leads_empty_for_admin_with_no_leads():
    client, headers = _admin_client(802)
    with client:
        response = client.get("/api/admin/leads", headers=headers)
    assert response.status_code == 200
    assert response.json() == []


def test_purchases_empty_for_admin_with_no_purchases():
    client, headers = _admin_client(803)
    with client:
        response = client.get("/api/admin/purchases", headers=headers)
    assert response.status_code == 200
    assert response.json() == []


def test_users_lists_self_after_first_request():
    # current_client() создаёт/трогает User-запись на каждый запрос — админ
    # уже появится в /api/admin/users после одного собственного запроса.
    client, headers = _admin_client(804)
    with client:
        response = client.get("/api/admin/users", headers=headers)
    assert response.status_code == 200
    users = response.json()
    assert len(users) == 1
    assert users[0]["tg_id"] == 804
    assert users[0]["checkpoint"] == "new"


def test_purchases_rejected_for_non_admin():
    client, headers = _client(805)
    with client:
        response = client.get("/api/admin/purchases", headers=headers)
    assert response.status_code == 403


def test_users_rejected_for_non_admin():
    client, headers = _client(806)
    with client:
        response = client.get("/api/admin/users", headers=headers)
    assert response.status_code == 403


def test_leads_returns_joined_user_fields_for_real_lead():
    # Same tg_id acts as the lead AND the admin checking the list — expected,
    # since owner_chat_id grants admin access to that tg_id regardless of what
    # else they do in the funnel. Walks the same consult/book -> consult/email
    # sequence as test_consult_email_valid_creates_lead_and_sets_idle in
    # test_funnel_api.py, which is the real code path that calls
    # db.crud.create_lead — this gives a real Lead row joined to a real User
    # row, exercising the `user.username if user else None` mapping in
    # _leads_out() with an actual (non-empty) result.
    client, headers = _admin_client_with_username(822, "lead_anna")
    with client:
        client.post("/api/funnel/consult/book", headers=headers)
        email_response = client.post(
            "/api/funnel/consult/email", headers=headers, json={"email": "someone@example.com"},
        )
        assert email_response.status_code == 200

        response = client.get("/api/admin/leads", headers=headers)
    assert response.status_code == 200
    leads = response.json()
    assert len(leads) == 1
    lead = leads[0]
    assert lead["tg_id"] == 822
    assert lead["username"] == "lead_anna"
    assert lead["email"] == "someone@example.com"
    assert lead["worked_at"] is None
    assert lead["created_at"] is not None


def test_purchases_returns_joined_user_fields_for_real_purchase(monkeypatch):
    # Same technique as test_buy_product_creates_purchase_and_returns_confirmation_url
    # in test_funnel_api.py: create_payment() would otherwise make a real
    # YooKassa HTTP call, so it's monkeypatched to a fake async implementation.
    # Everything happens inside one `with client:` block on one TestClient, so
    # the funnel purchase and the admin list read share the same sqlite
    # connection/event loop.
    async def fake_create_payment(**kwargs):
        return "yk-payment-999", "https://yookassa.ru/pay/yk-payment-999"

    monkeypatch.setattr("api.routers.funnel.create_payment", fake_create_payment)

    client, headers = _admin_client_with_username(823, "buyer_ivan")
    with client:
        client.post("/api/funnel/consent", headers=headers)
        for q, s in enumerate([2, 2, 0, 1, 0, 2, 1], start=1):
            client.post("/api/funnel/answers", headers=headers, json={"question_no": q, "score": s})
        buy_response = client.post(
            "/api/funnel/product/practicum/buy", headers=headers,
            json={"email": "buyer@example.com"},
        )
        assert buy_response.status_code == 200

        response = client.get("/api/admin/purchases", headers=headers)
    assert response.status_code == 200
    purchases = response.json()
    assert len(purchases) == 1
    purchase = purchases[0]
    assert purchase["tg_id"] == 823
    assert purchase["username"] == "buyer_ivan"
    assert purchase["product"] == "practicum"
    assert purchase["amount_rub"] == 2990
    assert purchase["status"] == "pending"  # no webhook fired -> still pending
    assert purchase["paid_at"] is None
    assert purchase["delivered_at"] is None
    assert isinstance(purchase["id"], int)


def test_leads_export_returns_xlsx_with_header_row():
    client, headers = _admin_client(807)
    with client:
        response = client.get("/api/admin/leads/export", headers=headers)
    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert [cell.value for cell in ws[1]] == ["tg_id", "username", "email", "worked", "created_at"]


def test_purchases_export_returns_xlsx_with_header_row():
    client, headers = _admin_client(808)
    with client:
        response = client.get("/api/admin/purchases/export", headers=headers)
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert [cell.value for cell in ws[1]] == [
        "id", "tg_id", "username", "product", "amount_rub", "status", "paid_at", "delivered_at",
    ]


def test_users_export_returns_xlsx_with_header_row():
    client, headers = _admin_client(809)
    with client:
        response = client.get("/api/admin/users/export", headers=headers)
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert [cell.value for cell in ws[1]] == [
        "tg_id", "username", "first_name", "checkpoint", "result_type", "test_attempt", "created_at",
    ]


def test_users_export_escapes_formula_injection_in_first_name():
    # CWE-1236 / OWASP CSV-Injection: Telegram first_name has no charset
    # restriction (unlike username), so a user can set it to a string that
    # Excel/LibreOffice/Google Sheets interprets as a formula when the admin
    # opens the exported .xlsx. _xlsx_response() must neutralize this by
    # prefixing any cell value starting with =, +, -, or @ with a leading
    # single-quote, which openpyxl (and Excel) treat as forced-text.
    #
    # Verified empirically (see task-7-report.md) that openpyxl preserves the
    # leading "'" verbatim on write+read-back as a plain string cell (not a
    # formula) -- so we assert on that literal round-tripped value.
    payload = "=SUM(A1:A10)"
    client, headers = _admin_client_with_first_name(825, payload)
    with client:
        response = client.get("/api/admin/users/export", headers=headers)
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    first_name_col = header.index("first_name")
    data_row = [cell.value for cell in ws[2]]
    assert data_row[first_name_col] == "'" + payload


def test_leads_export_rejected_for_non_admin():
    client, headers = _client(810)
    with client:
        response = client.get("/api/admin/leads/export", headers=headers)
    assert response.status_code == 403


def test_toggle_lead_worked_flips_status_and_back():
    client, headers = _admin_client_with_username(830, "lead_toggle")
    with client:
        client.post("/api/funnel/consult/book", headers=headers)
        client.post(
            "/api/funnel/consult/email", headers=headers, json={"email": "toggle@example.com"},
        )

        first = client.post("/api/admin/leads/830/worked", headers=headers)
        assert first.status_code == 200
        assert first.json()["worked_at"] is not None

        second = client.post("/api/admin/leads/830/worked", headers=headers)
        assert second.status_code == 200
        assert second.json()["worked_at"] is None


def test_toggle_lead_worked_404_for_unknown_lead():
    client, headers = _admin_client(831)
    with client:
        response = client.post("/api/admin/leads/999999/worked", headers=headers)
    assert response.status_code == 404


def test_toggle_lead_worked_rejected_for_non_admin():
    client, headers = _client(832)
    with client:
        response = client.post("/api/admin/leads/832/worked", headers=headers)
    assert response.status_code == 403


